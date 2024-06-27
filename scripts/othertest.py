import time
import json
import threading
from fake_useragent import UserAgent
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from threading import Thread
from queue import Queue

def fetch_urlvoid_data(domain, retries=3, timeout=60000):
    ua = UserAgent()
    attempt = 0
    while attempt < retries:
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)  # Headless mode for faster execution
                context = browser.new_context(
                    user_agent=ua.random,
                    viewport={'width': 1280, 'height': 800},
                    java_script_enabled=True,
                )
                page = context.new_page()
                print(f"Thread {threading.current_thread().name} - Fetching {domain}, attempt {attempt + 1}")
                urlvoid_url = f'https://www.urlvoid.com/scan/{domain}/'
                page.goto(urlvoid_url, timeout=timeout)

                # Wait for the score selector and get the score text
                page.wait_for_selector('.label.label-success', timeout=30000)
                score_element = page.query_selector('.label.label-success')
                score_text = score_element.inner_text() if score_element else 'N/A'
                
                browser.close()
                return domain, score_text
        except Exception as e:
            print(f"Thread {threading.current_thread().name} - Error fetching {domain}: {e}. Retrying...")
            attempt += 1
            time.sleep(2)  # Wait before retrying
    return domain, 'N/A'

def worker(queue, ws, wb, urls_scores, file_path):
    while not queue.empty():
        domain = queue.get()
        if domain is None:
            break
        print(f"Thread {threading.current_thread().name} - Working on {domain}")
        domain, score = fetch_urlvoid_data(domain)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == domain:
                    ws.cell(row=cell.row, column=2, value=score)
                    urls_scores[domain] = score
                    print(f"Thread {threading.current_thread().name} - Updated {domain} with score {score}")
                    wb.save(file_path)
        queue.task_done()
        print(f"Thread {threading.current_thread().name} - Finished {domain}")
        time.sleep(1)  # Add delay to prevent being blocked by the website

def update_excel_and_generate_json(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Assuming URLs are in column A and scores will be written in column B
    urls = [cell.value for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1) for cell in row]
    queue = Queue()
    urls_scores = {}

    # Populate the queue with URLs
    for url in urls:
        queue.put(url)

    # Create threads
    threads = []
    for i in range(10):  # 10 threads
        thread = Thread(target=worker, args=(queue, ws, wb, urls_scores, file_path), name=f"Thread-{i+1}")
        thread.start()
        threads.append(thread)

    # Wait for all threads to finish
    for thread in threads:
        thread.join()

    # Generate JSON output
    with open('url_scores.json', 'w') as json_file:
        json.dump(urls_scores, json_file, indent=4)
    print("JSON file created successfully.")

# Example usage
excel_file_path = 'links.xlsx'  # Replace with your Excel file path
update_excel_and_generate_json(excel_file_path)
